import csv
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook

MAX_PARSED_ROWS = 500


def _normalize_value(value):
    if value is None:
        return ""
    if isinstance(value, (int, float, bool, str)):
        return value
    return str(value)


def _parse_csv(content: bytes) -> tuple[list[str], list[dict], str]:
    text = content.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(StringIO(text))
    columns = [col for col in (reader.fieldnames or []) if col]

    rows: list[dict] = []
    for idx, row in enumerate(reader):
        if idx >= MAX_PARSED_ROWS:
            break
        parsed = {key: _normalize_value(row.get(key)) for key in columns}
        rows.append(parsed)

    status = "ready" if rows else "empty"
    return columns, rows, status


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict], str]:
    workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active

    iterator = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return [], [], "empty"

    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    columns = [header for header in headers if header]
    if not columns:
        return [], [], "empty"

    rows: list[dict] = []
    for idx, raw in enumerate(iterator):
        if idx >= MAX_PARSED_ROWS:
            break
        parsed = {}
        for index, key in enumerate(headers):
            if not key:
                continue
            value = raw[index] if index < len(raw) else ""
            parsed[key] = _normalize_value(value)
        rows.append(parsed)

    status = "ready" if rows else "empty"
    return columns, rows, status


def parse_tabular_document(filename: str, content: bytes) -> dict:
    extension = Path(filename or "").suffix.lower()

    try:
        if extension == ".csv":
            columns, rows, status = _parse_csv(content)
        elif extension in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            columns, rows, status = _parse_xlsx(content)
        else:
            return {
                "file_type": extension.replace(".", "") or "unknown",
                "parse_status": "unsupported",
                "parsed_columns": [],
                "parsed_rows": [],
                "row_count": 0,
            }

        return {
            "file_type": extension.replace(".", "") or "unknown",
            "parse_status": status,
            "parsed_columns": columns,
            "parsed_rows": rows,
            "row_count": len(rows),
        }
    except Exception:
        return {
            "file_type": extension.replace(".", "") or "unknown",
            "parse_status": "failed",
            "parsed_columns": [],
            "parsed_rows": [],
            "row_count": 0,
        }
